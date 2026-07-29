#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Apply manual-review candidate semantics back to MACSLU JSONL files.

Expected manual-review sheet columns:
  - split
  - text_id
  - 修改後

Rows are applied only when 修改後 starts with:
  候選補標，需人工確認：

The JSON list after the Chinese colon is written to the original JSONL field:
  obj["semantics"]

Because semantics is also embedded inside obj["text"], this script also rebuilds:
  language None<asr_text>{"asr_text": ..., "semantics": "[...]"}
"""

from __future__ import annotations

import argparse
import json
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl"
    ) from exc


CANDIDATE_PREFIX = "候選補標，需人工確認"
ASR_TAG = "<asr_text>"
REQUIRED_COLUMNS = {"split", "text_id", "修改後"}


CandidateMap = Dict[str, Dict[str, List[Dict[str, Any]]]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply manual-review candidate semantics to train/dev/test JSONL files."
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        required=True,
        help="Dataset directory containing train.jsonl, dev.jsonl, test.jsonl.",
    )
    parser.add_argument(
        "--review-xlsx",
        type=Path,
        required=True,
        help="Manual review xlsx file.",
    )
    parser.add_argument(
        "--sheet-name",
        default="v1",
        help="Sheet name in the manual review xlsx. Default: v1.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help=(
            "Directory for fixed JSONL files. If omitted, writes to "
            "<data-dir>_manual_fixed unless --inplace is used."
        ),
    )
    parser.add_argument(
        "--inplace",
        action="store_true",
        help="Overwrite JSONL files in --data-dir. A .bak file will be created once.",
    )
    parser.add_argument(
        "--splits",
        nargs="+",
        default=["train", "dev", "test"],
        help="Splits to process. Default: train dev test.",
    )
    parser.add_argument(
        "--copy-root-files",
        action="store_true",
        help=(
            "When using --output-dir, also copy root-level non-JSONL files "
            "such as labels.txt and semantic_fixed.txt."
        ),
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Raise an error if candidate text_id is not found in the target JSONL.",
    )
    return parser.parse_args()


def extract_candidate_semantics(cell_value: Any) -> List[Dict[str, Any]] | None:
    """Return parsed candidate semantics if the cell is an applicable candidate row."""
    if cell_value is None:
        return None

    text = str(cell_value).strip()
    if not text.startswith(CANDIDATE_PREFIX):
        return None

    rest = text[len(CANDIDATE_PREFIX) :].strip()
    if rest.startswith("：") or rest.startswith(":"):
        rest = rest[1:].strip()
    else:
        raise ValueError(f"Candidate cell has no colon after prefix: {text[:120]}")

    try:
        semantics = json.loads(rest)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Cannot parse candidate JSON: {rest[:200]}") from exc

    if isinstance(semantics, dict):
        semantics = [semantics]
    if not isinstance(semantics, list):
        raise ValueError(f"Candidate semantics must be a JSON list, got {type(semantics)}")
    if not all(isinstance(frame, dict) for frame in semantics):
        raise ValueError("Every semantic frame must be a JSON object.")

    return semantics


def load_candidates(review_xlsx: Path, sheet_name: str) -> CandidateMap:
    wb = load_workbook(review_xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col = {name: idx for idx, name in enumerate(headers)}
    missing_cols = REQUIRED_COLUMNS - set(col)
    if missing_cols:
        raise ValueError(f"Missing required columns in sheet '{sheet_name}': {sorted(missing_cols)}")

    candidates: CandidateMap = defaultdict(dict)
    duplicate_keys: List[Tuple[str, str, int]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        parsed = extract_candidate_semantics(row[col["修改後"]])
        if parsed is None:
            continue

        split = str(row[col["split"]]).strip()
        text_id = str(row[col["text_id"]]).strip()
        if not split or not text_id:
            raise ValueError(f"Row {row_idx}: split/text_id is empty.")

        if text_id in candidates[split]:
            duplicate_keys.append((split, text_id, row_idx))
            continue
        candidates[split][text_id] = parsed

    if duplicate_keys:
        preview = ", ".join(f"{s}/{tid}@row{r}" for s, tid, r in duplicate_keys[:10])
        raise ValueError(f"Duplicate candidate rows detected: {preview}")

    return {split: dict(items) for split, items in candidates.items()}


def dump_json_compact(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False)


def rebuild_text_field(old_text: Any, query: Any, new_semantics: List[Dict[str, Any]]) -> str:
    """
    Rebuild obj['text'] so its embedded semantics string matches obj['semantics'].

    Original style is usually:
      language None<asr_text>{"asr_text": "...", "semantics": "[...]"}
    """
    query_text = "" if query is None else str(query)
    semantics_as_string = dump_json_compact(new_semantics)

    prefix = f"language None{ASR_TAG}"
    payload: Dict[str, Any] = {}

    if isinstance(old_text, str) and ASR_TAG in old_text:
        left, right = old_text.split(ASR_TAG, 1)
        prefix = left + ASR_TAG
        right = right.strip()
        if right:
            try:
                loaded = json.loads(right)
                if isinstance(loaded, dict):
                    payload = loaded
            except json.JSONDecodeError:
                # If the old embedded JSON is broken, rebuild from scratch.
                payload = {}

    payload["asr_text"] = query_text
    payload["semantics"] = semantics_as_string
    return prefix + dump_json_compact(payload)


def read_jsonl(path: Path) -> Iterable[Tuple[int, Dict[str, Any]]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.rstrip("\n")
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path}:{line_no}: invalid JSONL line") from exc
            if not isinstance(obj, dict):
                raise ValueError(f"{path}:{line_no}: JSONL record must be an object")
            yield line_no, obj


def make_backup_once(path: Path) -> Path:
    backup_path = path.with_suffix(path.suffix + ".bak")
    if not backup_path.exists():
        shutil.copy2(path, backup_path)
    return backup_path


def process_split(
    split: str,
    input_path: Path,
    output_path: Path,
    split_candidates: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    total = 0
    updated = 0
    found_ids = set()
    update_examples = []

    with output_path.open("w", encoding="utf-8") as out_f:
        for line_no, obj in read_jsonl(input_path):
            total += 1
            text_id = str(obj.get("text_id", "")).strip()

            if text_id in split_candidates:
                new_semantics = split_candidates[text_id]
                old_semantics = obj.get("semantics")
                obj["semantics"] = new_semantics
                obj["text"] = rebuild_text_field(
                    old_text=obj.get("text", ""),
                    query=obj.get("query", ""),
                    new_semantics=new_semantics,
                )
                found_ids.add(text_id)
                updated += 1

                if len(update_examples) < 5:
                    update_examples.append(
                        {
                            "line_no": line_no,
                            "text_id": text_id,
                            "old_semantics": old_semantics,
                            "new_semantics": new_semantics,
                            "new_text": obj["text"],
                        }
                    )

            out_f.write(dump_json_compact(obj) + "\n")

    missing_ids = sorted(set(split_candidates) - found_ids)
    return {
        "split": split,
        "input_path": str(input_path),
        "output_path": str(output_path),
        "total_records": total,
        "candidate_rows": len(split_candidates),
        "updated_records": updated,
        "missing_text_ids": missing_ids,
        "examples": update_examples,
    }


def copy_root_files(data_dir: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for item in data_dir.iterdir():
        if item.is_file() and item.suffix != ".jsonl":
            shutil.copy2(item, output_dir / item.name)


def main() -> None:
    args = parse_args()

    data_dir = args.data_dir.resolve()
    review_xlsx = args.review_xlsx.resolve()
    if not data_dir.is_dir():
        raise FileNotFoundError(f"--data-dir not found: {data_dir}")
    if not review_xlsx.is_file():
        raise FileNotFoundError(f"--review-xlsx not found: {review_xlsx}")

    candidates = load_candidates(review_xlsx, args.sheet_name)
    selected_splits = set(args.splits)

    if args.inplace:
        output_dir = data_dir
    else:
        output_dir = args.output_dir.resolve() if args.output_dir else data_dir.with_name(data_dir.name + "_manual_fixed")
        output_dir.mkdir(parents=True, exist_ok=True)
        if args.copy_root_files:
            copy_root_files(data_dir, output_dir)

    reports = []
    missing_split_files = []

    for split in args.splits:
        input_path = data_dir / f"{split}.jsonl"
        if not input_path.exists():
            if candidates.get(split):
                missing_split_files.append(str(input_path))
            continue

        if args.inplace:
            make_backup_once(input_path)
            tmp_path = input_path.with_suffix(input_path.suffix + ".tmp")
            output_path = tmp_path
        else:
            output_path = output_dir / f"{split}.jsonl"

        report = process_split(
            split=split,
            input_path=input_path,
            output_path=output_path,
            split_candidates=candidates.get(split, {}),
        )

        if args.inplace:
            output_path.replace(input_path)
            report["output_path"] = str(input_path)

        reports.append(report)

    # Candidate rows whose split is not selected are intentionally ignored.
    ignored_splits = sorted(set(candidates) - selected_splits)

    summary = {
        "review_xlsx": str(review_xlsx),
        "sheet_name": args.sheet_name,
        "data_dir": str(data_dir),
        "output_dir": str(output_dir),
        "candidate_rows_total": sum(len(v) for v in candidates.values()),
        "candidate_rows_by_split": {k: len(v) for k, v in sorted(candidates.items())},
        "processed_splits": args.splits,
        "ignored_candidate_splits": ignored_splits,
        "missing_split_files": missing_split_files,
        "reports": reports,
    }

    all_missing_ids = [
        f"{r['split']}/{tid}"
        for r in reports
        for tid in r.get("missing_text_ids", [])
    ]
    if args.strict and (missing_split_files or all_missing_ids):
        raise RuntimeError(
            "Strict mode failed. Missing split files or candidate text_ids were found. "
            f"missing_split_files={missing_split_files}, missing_text_ids={all_missing_ids[:20]}"
        )

    report_dir = output_dir / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / "apply_manual_review_candidates_report.json"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    #print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"\nReport saved to: {report_path}")


if __name__ == "__main__":
    main()
